import warnings
import pandas as pd
from openpyxl import load_workbook
from getLTP import get_ltp

# # Get the price of the list of shares in a excel file
# workbook = Workbook()
# sheet = workbook.active
# sheet["A1"] = "Ticker"
# sheet["B1"] = "Price"
# for i in range(len(shares)):
#     nameCell = "A" + str(i + 2)
#     priceCell = "B" + str(i + 2)
#     sheet[nameCell] = shares[i]
#     sheet[priceCell] = float(get_ltp(shares[i]))
#     print(shares[i] + " : " + get_ltp(shares[i]))
#
# workbook.save(filename="hello_world.xlsx")

filename = 'C:\\Users\\linani\\Desktop\\Investments.xlsx'  # Actual file
# filename = "Investments.xlsx" #Test file
warnings.simplefilter("ignore", UserWarning)
excelfile = load_workbook(filename=filename)
# print(str(excelfile.sheetnames))
stock_sheet = excelfile['Stocks']
StockSymbolColumn = "B"
currentStockPriceColumn = "F"
starting = 2
df = pd.read_excel("C:\\Users\\linani\\Desktop\\Investments.xlsx", sheet_name='Stocks')
count = int(df[df['Name'] == 'Number of scrips'].index[0])

for i in range(starting, starting + count):
    stockSymbolCell = StockSymbolColumn + str(i)
    stockLTPCell = currentStockPriceColumn + str(i)
    print(str(stock_sheet[stockSymbolCell].value.upper()) + " : " + str(round(float(get_ltp(str(stock_sheet[stockSymbolCell].value.upper())))- float(stock_sheet[stockLTPCell].value),3))+ " : " + str(
        get_ltp(str(stock_sheet[stockSymbolCell].value.upper()))))
    stock_sheet[stockLTPCell] = float(get_ltp(str(stock_sheet[stockSymbolCell].value.upper())))

excelfile.save(filename=filename)
